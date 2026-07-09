"""
台帳同期スクリプト
ledger.xlsx を ledger.csv に変換して GitHub にプッシュする
"""
import subprocess
import sys
import os
from openpyxl import load_workbook

WORK_DIR  = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(WORK_DIR, "ledger.xlsx")
CSV_FILE   = os.path.join(WORK_DIR, "ledger.csv")
SHEET_NAME = "機種切替台帳"

os.chdir(WORK_DIR)

print("=== 台帳同期スクリプト ===")

# --- Excel → CSV 変換 ---
print(f"\n[1] Excelを読み込み中: {EXCEL_FILE}")
try:
    wb = load_workbook(EXCEL_FILE, data_only=True)
except FileNotFoundError:
    print(f"エラー: {EXCEL_FILE} が見つかりません")
    sys.exit(1)

ws = wb[SHEET_NAME]

rows = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if all(v is None for v in row):
        continue
    cells = []
    for v in row:
        if v is None:
            cells.append("")
        else:
            s = str(v).strip()
            # 日付型の場合 YYYY/MM/DD に整形
            if hasattr(v, 'strftime'):
                s = v.strftime("%Y/%m/%d")
            # カンマ含む場合はクォート
            if ',' in s or '"' in s:
                s = '"' + s.replace('"', '""') + '"'
            cells.append(s)
    rows.append(",".join(cells))

with open(CSV_FILE, "w", encoding="utf-8-sig", newline="\n") as f:
    f.write("\n".join(rows) + "\n")

print(f"    → CSV作成完了: {len(rows)-1} 件（ヘッダー除く）")

# --- Git push ---
print("\n[2] GitHubへアップロード中...")
try:
    subprocess.run(["git", "add", "ledger.csv", "index.html"], check=True)
    result = subprocess.run(
        ["git", "diff", "--cached", "--quiet"],
        capture_output=True
    )
    if result.returncode == 0:
        print("    → 変更なし。アップロード不要です。")
    else:
        subprocess.run(["git", "commit", "-m", "Update ledger"], check=True)
        subprocess.run(["git", "push"], check=True)
        print("    → GitHubへのアップロード完了！")
        print("\n数分後に以下のURLに反映されます:")
        print("    https://y1kinomura.github.io/sn-decoder/")
except subprocess.CalledProcessError as e:
    print(f"エラー: {e}")
    sys.exit(1)

print("\n=== 完了 ===")
